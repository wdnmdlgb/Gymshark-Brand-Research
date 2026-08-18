import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

wb = Workbook()

# 样式定义
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_alignment = Alignment(wrap_text=True, vertical='top')

def style_sheet(ws, df, title):
    ws.title = title
    # 写入标题行
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            cell.alignment = wrap_alignment
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # 列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[column].width = adjusted_width
    ws.row_dimensions[1].height = 30

# ========== Sheet1: 品牌基础信息对比 ==========
df1 = pd.DataFrame({
    '对比维度': ['创立时间', '总部', '创始人', '2024财年营收', '估值', '核心定位', '目标人群', '全球市场覆盖', '社媒粉丝总量', '品牌心智份额(健身训练类)'],
    'Gymshark': ['2012年', '英国 Solihull', 'Ben Francis(19岁创立)', '£607.3m(约7.5亿美元)', '约$1.45B', '健身服饰DTC品牌+社区文化', '16-30岁健身爱好者, 男女均衡', '200+国家, 14个线上商店', '1800万+', '36%(超过Under Armour的35%)'],
    'Lululemon': ['1998年', '加拿大温哥华', 'Chip Wilson', '约$100B(2024)', '约$50B+', '高端瑜伽运动生活方式品牌', '25-45岁高收入女性为主', '全球600+门店', '约2000万+', '瑜伽品类领先, 综合运动服饰第二梯队'],
    'Alphalete': ['2015年', '美国德州', 'Christian Guzman', '未公开(估$200-300m)', '未公开', '健身服饰DTC品牌, 美学导向', '18-30岁健身男性为主', '全球线上为主', '约500万+', '小众但忠诚度高']
})
ws1 = wb.active
style_sheet(ws1, df1, "品牌基础对比")

# ========== Sheet2: 产品线与价格对比 ==========
df2 = pd.DataFrame({
    '品类': ['女款Leggings', '男款训练短裤', '运动背心/Tank', '连帽衫/Hoodie', '运动内衣', '配饰(腰带/护腕等)'],
    'Gymshark代表系列': ['Vital / Energy+ / Flex', 'Power 5" / Vital 7" / Tactical', 'Everyday Racer Back / Essential', 'Lightweight / Fleece-lined', 'Essential / Whitney Simmons联名', 'lifting belts, straps, bottles'],
    'Gymshark价格(USD)': ['$55-$75', '$45-$65', '$30-$40', '$89-$115', '$35-$50', '$15-$45'],
    'Lululemon代表系列': ['Align / Wunder Train', 'Pace Breaker / T.H.E.', 'Align Tank / Ebb to Street', 'Scuba / Relaxed Fit', 'Energy / Free to Be', 'Everywhere Belt Bag等'],
    'Lululemon价格(USD)': ['$98-$148', '$68-$88', '$58-$78', '$118-$168', '$52-$68', '$38-$78'],
    'Alphalete代表系列': ['Halo / Pulse Kinetic / Amplify', 'Premium / Aero / Element', 'Aero Tank / Stratus', 'Elite / Premium', 'Aero / Pulse', 'lifting accessories'],
    'Alphalete价格(USD)': ['$60-$80', '$50-$70', '$38-$52', '$95-$130', '$42-$58', '$20-$50']
})
ws2 = wb.create_sheet()
style_sheet(ws2, df2, "产品线与价格对比")

# ========== Sheet3: 营销打法对比 ==========
df3 = pd.DataFrame({
    '营销维度': ['核心渠道', 'KOL策略', '标志性活动', '内容风格', '付费模式', '社区运营', '线下体验'],
    'Gymshark': ['TikTok(69.2%) + Instagram(30.5%) + YouTube(0.4%)', '微影响者为主+头部运动员持股(Chris Bumstead), 联盟营销按效果付费', 'Gymshark66挑战: 110万+帖子, 6570万TikTok互动', '真实、励志、训练日常, 不追求完美', 'TikTok Spark Ads + Instagram Reels为主', 'Gymshark Family社区, 运动员大使体系, 用户UGC', '伦敦/曼彻斯特/阿姆斯特丹/迪拜/纽约旗舰店, 德国店中店'],
    'Lululemon': ['Instagram + 线下门店活动 + 品牌大使', '高端瑜伽导师+社区KOL, 长期合作', 'Sweat Collective会员计划, 全球瑜伽日活动', '生活方式、高端、正念、社群感', '品牌广告+门店体验为主, 社媒投放较克制', '门店瑜伽课, 社区活动, 会员专属', '600+直营门店, 强线下体验'],
    'Alphalete': ['Instagram + YouTube + TikTok', '创始人Christian Guzman个人IP驱动, 健身网红合作', 'Alphalete Gym线下健身房, 品牌联名', '硬核健身、美学、肌肉展示', 'YouTube长视频+Instagram种草', 'YouTube社群, 线下健身房', '美国德州Alphalete Gym旗舰店']
})
ws3 = wb.create_sheet()
style_sheet(ws3, df3, "营销打法对比")

# ========== Sheet4: 欧洲本地化布局 ==========
df4 = pd.DataFrame({
    '本地化维度': ['欧洲线上商店', '欧洲线下门店', '本地支付支持', '本地语言支持', '学生折扣', '欧洲仓储/物流', '欧洲市场策略'],
    'Gymshark': ['英国/德国/法国/荷兰/西班牙/意大利/爱尔兰等14个', '阿姆斯特丹旗舰店(2025.4); 德国Engelhorn+Breuninger店中店(2026.2)', 'Klarna(德/法/荷/北欧), iDeal(荷兰), 信用卡/PayPal', '英/德/法/西/意/荷等多语言站点', 'UNiDAYS覆盖欧洲多国', '欧洲本地仓储, 支持快速配送', 'DTC为主, 2025起加速线下渗透, 首站荷兰+德国'],
    'Lululemon': ['全欧洲线上', '欧洲约50+门店(伦敦/巴黎/柏林/阿姆斯特丹等)', '本地支付+信用卡', '多语言', '部分国家有', '欧洲配送中心', '高端商场选址, 门店瑜伽课体验'],
    'Alphalete': ['欧洲线上配送', '无欧洲线下门店', '信用卡/PayPal为主', '英文为主', '无', '从美国发货, 配送较慢', '纯DTC, 未做欧洲本地化深度运营']
})
ws4 = wb.create_sheet()
style_sheet(ws4, df4, "欧洲本地化对比")

# ========== Sheet5: APP产品对比 ==========
df5 = pd.DataFrame({
    'APP维度': ['APP名称', '核心功能', '收费模式', '课程数量', '用户评分', '与品牌协同', '差异化亮点'],
    'Gymshark': ['Gymshark Training and Fitness', '450+免费训练课程, 自定义计划, 训练记录, 休息计时器, #Gymshark66挑战联动', '100%免费, 无广告', '450+', 'App Store约4.7/5(购物APP); 训练APP评价正面', '免费APP引流→品牌社区→服饰购买, 与Gymshark66活动深度绑定', '完全免费, 运动员课程内容, 品牌社区联动'],
    'Lululemon': ['Lululemon Studio(原Mirror)', '线上健身课程, 瑜伽/力量/有氧, 直播课', '订阅制$39/月(原Mirror硬件+$39/月)', '10000+', '约4.5/5', '高端会员生态, 硬件+内容+服饰闭环', '高端直播课, 名师资源, 硬件联动'],
    'Alphalete': ['无独立训练APP', '—', '—', '—', '—', '依赖创始人YouTube频道内容', '—']
})
ws5 = wb.create_sheet()
style_sheet(ws5, df5, "APP产品对比")

# ========== Sheet6: Gymshark爆款单品分析 ==========
df6 = pd.DataFrame({
    '爆款单品': ['Vital Seamless Leggings', 'Power 5" Shorts(蓝标)', 'Everyday Racer Back Tank', 'Flex Training Shorts', 'Whitney Simmons联名系列'],
    '品类': ['女款紧身裤', '男款训练短裤', '女款运动背心', '男款训练短裤', '女款联名系列'],
    '价格(USD)': ['$55-$65', '$55-$62', '$38', '$55-$65', '$60-$80'],
    '爆款原因': ['无缝针织技术, 提臀剪裁, 多色系, 社交媒体高频露出', '标志性蓝标, 修身剪裁, 健身圈文化符号, 得物炒至¥1289', '基础百搭, 带胸垫, 性价比高, 日常+训练两用', '轻量化面料, 适合HIIT/瑜伽, 多场景穿着', '网红Whitney Simmons联名, 粉丝经济, 限定发售制造稀缺'],
    '用户口碑关键词': ['squat-proof, flattering, comfortable', 'iconic, fitted, must-have', 'comfy, versatile, great value', 'lightweight, breathable, mobility', 'aesthetic, limited, collectible']
})
ws6 = wb.create_sheet()
style_sheet(ws6, df6, "Gymshark爆款单品分析")

# 保存
output_path = r"C:\Users\87090\Doubao\chats\2026-08-18\new-chat-1\项目1-Gymshark品牌研究\data\Gymshark竞品对比分析表.xlsx"
wb.save(output_path)
print(f"Excel已生成: {output_path}")
print(f"共{len(wb.sheetnames)}个Sheet: {wb.sheetnames}")
